"""
============================================================
Compile per-participant behavioural-analysis Excel
============================================================

Joins `extracted_survey_data.csv` (per participant x image) with the master
`Foodpictures_information_canonical.csv` (per image) and writes a single
Excel sheet for behavioural analysis.

DESIGN PRINCIPLE
----------------
This script is deliberately AGNOSTIC to the AI rating schema in canonical.

- AI code (e.g. LLM_image_QC.py for `aware_ai_*`, LLM_image_blind_ratings.py
  for `blind_ai_*`) writes its columns directly into canonical.
- Human-rating aggregation (update_canonical_human_means.py) writes its
  columns directly into canonical.
- This script only joins. Every canonical column passes through unchanged
  with its original name. New AI rating sets added to canonical tomorrow
  appear in this Excel automatically -- no edits required here.

SURVEY-SIDE RENAMES
-------------------
The seven per-participant taste columns are renamed (Sweet -> SweetHuman,
etc.) to keep them visually distinct from any per-image columns that share
similar names. FoodName -> Food Label_Human for the same reason. Nothing
else is renamed.

INPUTS
------
- ./extracted_survey_data.csv
- ../data/Foodpictures_information_canonical.csv

OUTPUT
------
- ./ai_vs_human_COMPILED.xlsx
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SCRIPT_DIR = Path(__file__).resolve().parent
SURVEY_CSV = SCRIPT_DIR / "extracted_survey_data.csv"
CANONICAL_CSV = SCRIPT_DIR.parent / "data" / "Foodpictures_information_canonical.csv"
OUTPUT_XLSX = SCRIPT_DIR / "ai_vs_human_COMPILED.xlsx"

# Only these survey columns get renamed (collision avoidance with canonical)
SURVEY_RENAME = {
    "Sweet":    "SweetHuman",
    "Salty":    "SaltyHuman",
    "Sour":     "SourHuman",
    "Bitter":   "BitterHuman",
    "Umami":    "UmamiHuman",
    "Fatty":    "FattyHuman",
    "Spicy":    "SpicyHuman",
    "FoodName": "Food Label_Human",
}

# Header fill colours by source/family
SURVEY_FILL    = "FCF3CF"  # pale yellow  - per-participant survey columns
HUMAN_AGG_FILL = "FAD7A0"  # peach        - canonical human_* (per-image means)
AWARE_AI_FILL  = "D7BDE2"  # lavender     - canonical aware_ai_*
BLIND_AI_FILL  = "AED6F1"  # pale blue    - canonical blind_ai_*
LL_FILL        = "D5F5E3"  # pale green   - canonical ll_* (low-level visual)
DL_FILL        = "D1F2EB"  # pale teal    - canonical dl_* (deep-learning)
OTHER_FILL     = "EAEDED"  # neutral grey - image metadata, QC, anything else


def fill_for(col_name: str, survey_cols: set) -> str:
    if col_name in survey_cols:
        return SURVEY_FILL
    if col_name.startswith("human_"):
        return HUMAN_AGG_FILL
    if col_name.startswith("aware_ai_"):
        return AWARE_AI_FILL
    if col_name.startswith("blind_ai_"):
        return BLIND_AI_FILL
    if col_name.startswith("ll_"):
        return LL_FILL
    if col_name.startswith("dl_"):
        return DL_FILL
    return OTHER_FILL


def main():
    if not SURVEY_CSV.exists():
        sys.exit(f"[ERROR] Survey extraction not found: {SURVEY_CSV}\n"
                 f"        Run survey_analysis.py first.")
    if not CANONICAL_CSV.exists():
        sys.exit(f"[ERROR] Canonical CSV not found: {CANONICAL_CSV}")

    print(f"[INFO] Loading survey:    {SURVEY_CSV.name}")
    survey = pd.read_csv(SURVEY_CSV)
    print(f"        {len(survey):,} rows, {len(survey.columns)} columns")

    print(f"[INFO] Loading canonical: {CANONICAL_CSV.name}")
    canonical = pd.read_csv(CANONICAL_CSV)
    print(f"        {len(canonical):,} images, {len(canonical.columns)} columns")

    if "filename" not in canonical.columns:
        sys.exit("[ERROR] Canonical CSV has no `filename` column - cannot match.")
    if "ImageName" not in survey.columns:
        sys.exit("[ERROR] Survey extraction has no `ImageName` column.")

    # ---- Survey-side renames (collision avoidance only) ----
    survey = survey.rename(columns=SURVEY_RENAME)
    survey_cols = list(survey.columns)

    # ---- Join: every survey row gains all canonical columns for its image ----
    df = survey.merge(
        canonical,
        left_on="ImageName",
        right_on="filename",
        how="left",
    ).drop(columns=["filename"], errors="ignore")

    # ---- Column ordering: survey-side first, then everything from canonical ----
    canonical_only = [c for c in df.columns if c not in survey_cols]
    df = df[[c for c in survey_cols if c in df.columns] + canonical_only]

    # ---- Diagnostics ----
    canonical_filenames = set(canonical["filename"].dropna())
    unmatched = sorted(set(survey["ImageName"].dropna()) - canonical_filenames)
    print(f"[INFO] Joined {len(df):,} survey rows with {len(canonical_only)} canonical columns")
    if unmatched:
        print(f"[WARN] {len(unmatched)} survey ImageName(s) had no match in canonical:")
        for n in unmatched[:10]:
            print(f"         {n}")
        if len(unmatched) > 10:
            print(f"         ... and {len(unmatched) - 10} more")

    # ---- Write Excel ----
    print(f"[INFO] Writing {OUTPUT_XLSX.name}...")
    df.to_excel(OUTPUT_XLSX, index=False)

    # ---- Header colour-coding ----
    survey_col_set = set(survey_cols)
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    for col_idx, col_name in enumerate(df.columns, start=1):
        colour = fill_for(col_name, survey_col_set)
        ws.cell(row=1, column=col_idx).fill = PatternFill(
            start_color=colour, end_color=colour, fill_type="solid"
        )
    wb.save(OUTPUT_XLSX)

    print(f"[DONE] {OUTPUT_XLSX.name}: {len(df):,} rows, {len(df.columns)} columns")
    print(f"        survey-side: {len(survey_cols)} cols (yellow)")
    print(f"        canonical:   {len(canonical_only)} cols "
          f"(peach=human_means, lavender=aware_ai, blue=blind_ai, green=ll, teal=dl, grey=other)")


if __name__ == "__main__":
    main()
