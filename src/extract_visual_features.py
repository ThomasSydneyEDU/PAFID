#!/usr/bin/env python3
"""
VISUAL METRICS EXTRACTION SCRIPT

DESCRIPTION
-----------
This script processes images and computes low-level (ll_) visual features 
matching the original FoodTriplet-Analysis baseline mathematical computations.

Usage:
  python src/extract_visual_features.py --stimuli-dir rendered_images/
"""

import os
import argparse
import cv2
import numpy as np
import pandas as pd
from tqdm import tqdm
from pathlib import Path
from sklearn.decomposition import PCA

HAS_OPENPYXL = True
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    HAS_OPENPYXL = False

EXPECTED_COLS = [
    "filename",
    "ll_mean_luminance", "ll_rms_contrast", 
    "ll_lab_L_mean", "ll_lab_L_std", "ll_lab_a_mean", "ll_lab_a_std", 
    "ll_lab_b_mean", "ll_lab_b_std", "ll_hsv_s_mean", "ll_edge_energy"
] + [f"ll_hog_pc{j+1:02d}" for j in range(10)]


def compute_legacy_ll_metrics(image_bgr):
    from skimage.color import rgb2gray, rgb2lab, rgb2hsv
    from skimage.filters import sobel
    from skimage.feature import hog

    # BGR to RGB
    rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    
    # Skimage float32 normalization
    rgb_float = rgb.astype(np.float32) / 255.0

    gray = rgb2gray(rgb_float).astype(np.float32)
    mean_lum = float(np.mean(gray))
    rms_contrast = float(np.std(gray))

    lab = rgb2lab(rgb_float).astype(np.float32)
    L = lab[..., 0]
    a = lab[..., 1]
    b = lab[..., 2]

    hsv = rgb2hsv(rgb_float).astype(np.float32)
    s = hsv[..., 1]

    edge_energy = float(np.mean(np.abs(sobel(gray))))

    hog_feat = hog(
        gray,
        orientations=9,
        pixels_per_cell=(16, 16),
        cells_per_block=(2, 2),
        block_norm="L2-Hys",
        feature_vector=True,
    ).astype(np.float32)

    return {
        "ll_mean_luminance": mean_lum,
        "ll_rms_contrast": rms_contrast,
        "ll_lab_L_mean": float(np.mean(L)),
        "ll_lab_L_std": float(np.std(L)),
        "ll_lab_a_mean": float(np.mean(a)),
        "ll_lab_a_std": float(np.std(a)),
        "ll_lab_b_mean": float(np.mean(b)),
        "ll_lab_b_std": float(np.std(b)),
        "ll_hsv_s_mean": float(np.mean(s)),
        "ll_edge_energy": edge_energy,
        "hog_raw": hog_feat
    }


def style_excel_columns(xlsx_path):
    if not HAS_OPENPYXL: return
    wb = load_workbook(xlsx_path)
    ws = wb.active
    groups = {
        "filename": (["filename"], "D9D9D9"),
        "legacy_ll": (["ll_mean_luminance", "ll_rms_contrast", "ll_lab_L_mean", "ll_lab_L_std", "ll_lab_a_mean", "ll_lab_a_std", "ll_lab_b_mean", "ll_lab_b_std", "ll_hsv_s_mean", "ll_edge_energy"] + [f"ll_hog_pc{j+1:02d}" for j in range(10)], "E2EFDA"),
    }
    header_row = 1
    headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row]) if cell.value is not None}
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for cols, color in groups.values():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for colname in cols:
            if colname not in headers: continue
            col_letter = get_column_letter(headers[colname])
            cell = ws[f"{col_letter}{header_row}"]
            cell.fill, cell.font = fill, Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    for name, idx in headers.items():
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(str(name)) + 4)
    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Extract low-level visual features from images.")
    parser.add_argument("--stimuli-dir", type=str, required=True, help="Folder containing images.")
    parser.add_argument("--output-csv", type=str, default=None, help="Path to save CSV output. Defaults to <stimuli-dir>/visual_metrics.csv")
    parser.add_argument("--merge-canonical", action="store_true", help="Merge results into data/Foodpictures_information_dynamic.csv")
    parser.add_argument("--overwrite", action="store_true",
                        help="Replace ll_ values for ALL rows when merging. Default is incremental: "
                             "only rows with missing ll_ values (i.e. new stimuli) are filled, preserving "
                             "the canonical baseline values for existing items.")
    args = parser.parse_args()

    stimuli_dir = Path(args.stimuli_dir).resolve()
    if not args.output_csv:
        output_csv = stimuli_dir / "visual_metrics.csv"
    else:
        output_csv = Path(args.output_csv)

    print(f"[INFO] Processing images in: {stimuli_dir}")
    files = sorted([f for f in os.listdir(stimuli_dir) if f.lower().endswith((".png",".jpg",".jpeg"))])
    print(f"[INFO] Found {len(files)} images")

    results = []
    for fname in tqdm(files):
        path = stimuli_dir / fname
        img = cv2.imread(str(path))
        if img is None: continue
        row = {c: np.nan for c in EXPECTED_COLS}; row["filename"] = fname

        try:
            ll_feats = compute_legacy_ll_metrics(img)
            for k, v in ll_feats.items():
                row[k] = v
        except Exception as e:
            print(f"Error computing legacy LL metrics on {fname}: {e}")

        results.append(row)

    # Perform PCA on HOG features across all images
    hog_raws = []
    valid_mask = []
    for r in results:
        if 'hog_raw' in r and isinstance(r['hog_raw'], np.ndarray):
            hog_raws.append(r['hog_raw'])
            valid_mask.append(True)
        else:
            hog_raws.append(np.zeros(1))
            valid_mask.append(False)

    if sum(valid_mask) > 5:
        X = np.array([h for m, h in zip(valid_mask, hog_raws) if m])
        from sklearn.preprocessing import StandardScaler
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        n_pcs = min(10, X_scaled.shape[0], X_scaled.shape[1])
        pca = PCA(n_components=n_pcs, random_state=0)
        X_pca = pca.fit_transform(X_scaled)
        
        pca_idx = 0
        for i, m in enumerate(valid_mask):
            for j in range(10):
                col_name = f"ll_hog_pc{j+1:02d}"
                if m and j < n_pcs:
                    results[i][col_name] = float(X_pca[pca_idx, j])
                else:
                    results[i][col_name] = np.nan
            if m:
                pca_idx += 1
    else:
        for i in range(len(results)):
            for j in range(10):
                results[i][f"ll_hog_pc{j+1:02d}"] = np.nan

    # Clean up raw HOG feature
    for r in results:
        if 'hog_raw' in r:
            del r['hog_raw']

    df_metrics = pd.DataFrame(results, columns=EXPECTED_COLS)
    df_metrics.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"[OK] Saved CSV: {output_csv}")

    if HAS_OPENPYXL:
        output_xlsx = output_csv.with_suffix(".xlsx")
        df_metrics.to_excel(output_xlsx, index=False, engine="openpyxl")
        style_excel_columns(output_xlsx)
        print(f"[OK] Saved styled Excel: {output_xlsx}")

    if args.merge_canonical:
        src_dir = Path(__file__).resolve().parent
        canonical_path = src_dir.parent / "data" / "Foodpictures_information_dynamic.csv"
        if canonical_path.exists():
            print(f"[INFO] Merging results into: {canonical_path}")
            try:
                # utf-8-sig: the dynamic CSV is written with a BOM by run_qc.py;
                # reading without it breaks the 'filename' column lookup (the BOM
                # becomes part of the first header) and the merge silently fails.
                df_can = pd.read_csv(canonical_path, encoding="utf-8-sig")
                if "filename" not in df_can.columns:
                    print(f"[ERROR] 'filename' column missing in {canonical_path}. Merge failed.")
                else:
                    ll_cols = [c for c in EXPECTED_COLS if c != 'filename']
                    df_metrics_dedup = df_metrics.drop_duplicates(subset="filename", keep="first")

                    if args.overwrite:
                        # Replace ll_ values for all rows.
                        df_can.drop(columns=[c for c in ll_cols if c in df_can.columns], inplace=True)
                        df_merged = df_can.merge(df_metrics_dedup, on="filename", how="left")
                        print(f"[INFO] --overwrite: replaced ll_ values for all rows.")
                    else:
                        # Incremental (default): only fill rows whose ll_ values are
                        # missing (new stimuli). Existing values — e.g. the canonical
                        # 350-item baseline — are preserved. NOTE: HOG PCs are
                        # PCA-based and depend on the image set they were fit on,
                        # so values from different runs are not in the same basis.
                        df_merged = df_can.copy()
                        for c in ll_cols:
                            if c not in df_merged.columns:
                                df_merged[c] = pd.NA
                        lookup = df_metrics_dedup.set_index("filename")
                        mask = df_merged[ll_cols[0]].isna()
                        n_fill = 0
                        for idx in df_merged.index[mask]:
                            fn = str(df_merged.at[idx, "filename"])
                            if fn in lookup.index:
                                for c in ll_cols:
                                    df_merged.at[idx, c] = lookup.at[fn, c]
                                n_fill += 1
                        print(f"[INFO] Incremental merge: filled ll_ values for {n_fill} new row(s); "
                              f"{int((~mask).sum())} existing row(s) preserved. Use --overwrite to recompute all.")

                    # Save back
                    df_merged.to_csv(canonical_path, index=False, encoding="utf-8-sig")
                    print("[OK] Merged results into canonical CSV.")
            except Exception as e:
                print(f"[ERROR] Could not merge into canonical: {e}")
        else:
            print(f"[WARN] Canonical CSV not found at {canonical_path}. Merge skipped.")

if __name__ == "__main__":
    main()
